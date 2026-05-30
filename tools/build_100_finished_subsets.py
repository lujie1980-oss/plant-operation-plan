import json
from pathlib import Path
from openpyxl import load_workbook, Workbook


def read_header(ws):
    for r in range(1, 6):
        vals = [ws.cell(r, c).value for c in range(1, 80)]
        if any(v not in (None, "") for v in vals):
            header = [str(v).strip() if v is not None else "" for v in vals]
            while header and header[-1] == "":
                header.pop()
            return r, header
    raise ValueError(f"No header found in sheet {ws.title}")


def index_map(header):
    return {name: i for i, name in enumerate(header)}


def get_cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def copy_filtered_rows(src_ws, dst_ws, header_row, header, keep_fn):
    for c, h in enumerate(header, start=1):
        dst_ws.cell(1, c).value = h
    out_r = 2
    for r in range(header_row + 1, src_ws.max_row + 1):
        row_vals = [src_ws.cell(r, c).value for c in range(1, len(header) + 1)]
        if all(v in (None, "") for v in row_vals):
            continue
        if not keep_fn(row_vals):
            continue
        for c, v in enumerate(row_vals, start=1):
            dst_ws.cell(out_r, c).value = v
        out_r += 1
    return out_r - 2


def main():
    root = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp")
    material_bom_file = root / "routing-bom.xlsx"   # this file is material BOM template
    routing_bom_file = root / "material-bom.xlsx"   # this file is routing BOM template

    subset_material = root / "subset-material-bom-100.xlsx"
    subset_routing = root / "subset-routing-bom-100.xlsx"

    wb_mat = load_workbook(material_bom_file, data_only=False)
    ws_mat = wb_mat[wb_mat.sheetnames[0]]
    mat_header_row, mat_header = read_header(ws_mat)
    mat_idx = index_map(mat_header)

    finished_col = None
    for k in ("成品料号", "成品编码"):
        if k in mat_idx:
            finished_col = mat_idx[k]
            break
    if finished_col is None:
        raise ValueError("Material BOM file missing 成品料号 column")

    finished_codes = []
    finished_set = set()
    for r in range(mat_header_row + 1, ws_mat.max_row + 1):
        v = get_cell_str(ws_mat.cell(r, finished_col + 1).value)
        if not v:
            continue
        if v in finished_set:
            continue
        finished_set.add(v)
        finished_codes.append(v)
        if len(finished_codes) >= 100:
            break

    top100 = set(finished_codes)

    out_mat = Workbook()
    out_ws_mat = out_mat.active
    out_ws_mat.title = ws_mat.title

    mat_rows = copy_filtered_rows(
        ws_mat,
        out_ws_mat,
        mat_header_row,
        mat_header,
        lambda row: get_cell_str(row[finished_col]) in top100,
    )

    if len(wb_mat.sheetnames) > 1:
        for extra in wb_mat.sheetnames[1:]:
            src = wb_mat[extra]
            dst = out_mat.create_sheet(extra)
            for r in range(1, src.max_row + 1):
                for c in range(1, src.max_column + 1):
                    dst.cell(r, c).value = src.cell(r, c).value

    out_mat.save(subset_material)

    wb_route = load_workbook(routing_bom_file, data_only=False)
    ws_route = wb_route[wb_route.sheetnames[0]]
    route_header_row, route_header = read_header(ws_route)
    route_idx = index_map(route_header)

    route_finished_col = route_idx.get("成品料号")
    route_material_col = route_idx.get("料号")
    if route_finished_col is None and route_material_col is None:
        raise ValueError("Routing BOM file missing 成品料号/料号 column")

    out_route = Workbook()
    out_ws_route = out_route.active
    out_ws_route.title = ws_route.title

    def keep_route(row):
        f = get_cell_str(row[route_finished_col]) if route_finished_col is not None else ""
        m = get_cell_str(row[route_material_col]) if route_material_col is not None else ""
        return (f in top100) or (m in top100)

    route_rows = copy_filtered_rows(
        ws_route,
        out_ws_route,
        route_header_row,
        route_header,
        keep_route,
    )

    if len(wb_route.sheetnames) > 1:
        for extra in wb_route.sheetnames[1:]:
            src = wb_route[extra]
            dst = out_route.create_sheet(extra)
            for r in range(1, src.max_row + 1):
                for c in range(1, src.max_column + 1):
                    dst.cell(r, c).value = src.cell(r, c).value

    out_route.save(subset_routing)

    print(json.dumps({
        "top100_count": len(top100),
        "material_bom_rows": mat_rows,
        "routing_bom_rows": route_rows,
        "subset_material_bom": str(subset_material),
        "subset_routing_bom": str(subset_routing),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
