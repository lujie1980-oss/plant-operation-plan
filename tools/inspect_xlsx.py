import os
import sys
import json

from openpyxl import load_workbook


def inspect(path: str) -> None:
    print(f"\n=== {path}")
    if not os.path.exists(path):
        print("MISSING")
        return
    wb = load_workbook(path, data_only=True)
    print("sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        sh = wb[name]

        header = None
        header_row = None
        for r in range(1, 6):
            row = [sh.cell(r, c).value for c in range(1, 51)]
            if any(v not in (None, "") for v in row):
                header = row
                header_row = r
                break
        if header is None:
            print(f" - {name}: empty")
            continue
        header = [str(v).strip() if v is not None else "" for v in header]
        while header and header[-1] == "":
            header.pop()
        print(f" - {name}: header_row={header_row} cols={len(header)}")
        print("   headers_json:", json.dumps(header, ensure_ascii=True))


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python tools/inspect_xlsx.py <file.xlsx> [more.xlsx...]")
        return 2
    for p in argv[1:]:
        inspect(p)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))

