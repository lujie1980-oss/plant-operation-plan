from openpyxl import Workbook
from pathlib import Path
p = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\one-row-material-bom.xlsx")
wb = Workbook()
ws = wb.active
headers = ["基地代码","BOM版本","BOM生效时间","BOM失效时间","成品料号","产品代码","组件代码","不计算齐套率","组件数量","组件损耗率","组件生效时间","组件失效时间"]
for i,h in enumerate(headers, start=1):
    ws.cell(1,i).value = h
row = ["TE","V1","2026-01-01","2026-12-31","FG-001","P-001","C-001","否",1,0.02,"2026-01-01","2026-12-31"]
for i,v in enumerate(row, start=1):
    ws.cell(2,i).value = v
wb.save(p)
print(str(p))
