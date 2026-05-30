"""Convert 盾安数据收集模板 → factory-demo.json + 匹配报告.

输出:
  - src/main/resources/sample-data/factory-dunan-demo.json
  - tools/dunan_match_report.json
"""
from __future__ import annotations

import json
import statistics
from collections import OrderedDict, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

EXTRACTED = Path(__file__).parent / "dunan_extracted.json"
OUTPUT = Path(__file__).parent.parent / "src" / "main" / "resources" / "sample-data" / "factory-dunan-demo.json"
OUTPUT_LITE = Path(__file__).parent.parent / "src" / "main" / "resources" / "sample-data" / "factory-dunan-demo-lite.json"
OUTPUT_EXCEL = Path(__file__).parent / "dunan-master-data.xlsx"
REPORT = Path(__file__).parent / "dunan_match_report.json"

CUSTOMER = "DUNAN"
AREA_DEFAULT = "2114"


def safe_float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def safe_int(v: Any, default: int = 0) -> int:
    if v is None:
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def parse_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_takt_seconds(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            if len(parts) == 3:
                h, m, sec = (int(parts[0]), int(parts[1]), int(float(parts[2])))
                return float(h * 3600 + m * 60 + sec)
            if len(parts) == 2:
                m, sec = int(parts[0]), int(float(parts[1]))
                return float(m * 60 + sec)
        except ValueError:
            return None
    return safe_float(s, 0) or None


def build_routing_index(data: dict[str, list[dict]]) -> tuple[
    dict[str, set[tuple[str, str | None]]],
    dict[tuple[str, str | None], list[dict]],
    dict[str, dict],
]:
    """material -> {(routing_code, step)}; (routing,step) -> equipment rows; resource meta."""
    proc_by_mat: dict[str, set[tuple[str, str | None]]] = defaultdict(set)
    for row in data["工艺"]:
        mat = row.get("产品代码")
        gy = row.get("工艺流程编号")
        step = row.get("工序步骤号")
        if mat and gy:
            proc_by_mat[mat].add((gy, str(step) if step is not None else None))

    equip_by_gy: dict[tuple[str, str | None], list[dict]] = defaultdict(list)
    resources_meta: dict[str, dict] = {}
    for row in data["生产设备"]:
        gy = row.get("工艺路线代码")
        step = row.get("工序步骤号")
        if not gy:
            continue
        key = (gy, str(step) if step is not None else None)
        wc_code = row.get("*设备代码")
        wc_name = row.get("*设备名称")
        line_code = row.get("*设备代码_2")
        line_name = row.get("*设备名称_2")
        op_code = row.get("*工序代码_2") or row.get("*工序代码")
        op_name = row.get("*工序名称")
        resource_id = line_name or line_code or wc_name or wc_code
        if not resource_id:
            continue
        equip_by_gy[key].append({
            "resourceId": resource_id,
            "wcCode": wc_code,
            "wcName": wc_name,
            "lineCode": line_code,
            "lineName": line_name,
            "operationCode": op_code,
            "operationName": op_name,
            "routingCode": gy,
            "step": key[1],
        })
        if resource_id not in resources_meta:
            area = row.get("*工序名称") or AREA_DEFAULT
            resources_meta[resource_id] = {
                "resourceId": resource_id,
                "areaId": str(row.get("*车间代码") or AREA_DEFAULT),
                "resourceGroup": row.get("设备组"),
                "bottleneck": False,
            }

    return proc_by_mat, equip_by_gy, resources_meta


def build_explicit_cpe_routes(data: dict[str, list[dict]]) -> list[dict]:
    routes: list[dict] = []
    for row in data["设备-产品生产关系"]:
        cpe = row.get("*产品代码")
        if not cpe or not str(cpe).upper().startswith("CPE"):
            continue
        resource_id = row.get("*设备名称_2") or row.get("*设备代码_2") or row.get("*设备名称") or row.get("*设备代码")
        if not resource_id:
            continue
        routes.append({
            "productCode": cpe,
            "resourceId": resource_id,
            "sequenceNo": safe_int(row.get("步骤号"), 1),
            "operationName": row.get("*工序类型"),
            "processTimeSeconds": parse_takt_seconds(row.get("生产节拍\n（单位：秒）")),
            "materialCode": row.get("*产品代码_2"),
            "routingCode": row.get("工艺路线代码"),
        })
    return routes


def resolve_order_product(
    mat: str | None,
    cpe_hint: str | None,
    proc_by_mat: dict[str, set[tuple[str, str | None]]],
    equip_by_gy: dict[tuple[str, str | None], list[dict]],
) -> tuple[str | None, list[dict], str | None]:
    """Returns (productCode, equipment hits, issue)."""
    if cpe_hint and str(cpe_hint).upper().startswith("CPE"):
        return cpe_hint, [], None
    if not mat:
        return None, [], "missing_product_code"
    keys = proc_by_mat.get(mat)
    if not keys:
        return mat, [], "no_routing_in_工艺"
    hits: list[dict] = []
    for key in keys:
        hits.extend(equip_by_gy.get(key, []))
    if not hits:
        return mat, [], "no_equipment_for_routing"
    return mat, hits, None


def write_master_data_excel(data: OrderedDict, path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name: str, headers: list[str], rows: list[list[Any]]) -> None:
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    sheet(
        "BOM",
        ["系统ID(留空=新增)", "BOM ID", "版本", "父产品", "组件产品", "用量", "关键件(是/否)"],
        [
            ["", "BOM-DUNAN", "V1", b["parentProductCode"], b["componentProductCode"], b["componentQty"], "是"]
            for b in data["bomComponents"]
        ],
    )
    sheet(
        "生产资源",
        ["系统ID(留空=新增)", "资源 ID", "资源组", "区域", "瓶颈(是/否)", "小时产能"],
        [
            ["", r["resourceId"], "", r["areaId"], "是" if r["bottleneck"] else "否", r["runRatePerHour"]]
            for r in data["resources"]
        ],
    )
    sheet(
        "产品工艺",
        ["系统ID(留空=新增)", "产品", "工序号", "工序名称", "资源", "换型(分钟)", "单件加工(秒)"],
        [
            [
                "",
                p["productCode"],
                p.get("sequenceNo"),
                p.get("operationName"),
                p["resourceId"],
                p.get("setupTimeMinutes"),
                p.get("processTimeSeconds"),
            ]
            for p in data["productResources"]
        ],
    )
    sheet(
        "产线",
        ["系统ID(留空=新增)", "产线 ID", "区域", "关联资源", "最小人数", "每班产能(分钟)"],
        [
            ["", ln["lineId"], ln["areaId"], ln["resourceId"], ln["lineMinHeadcount"], ln["lineCapacityPerShift"]]
            for ln in data["lines"]
        ],
    )
    wb.create_sheet("资源日历")
    wb["资源日历"].append(["系统ID(留空=新增)", "资源", "日期(yyyy-MM-dd)", "班次", "可用(分钟)", "不可用(分钟)"])
    wb.create_sheet("班次人员")
    wb["班次人员"].append(["系统ID(留空=新增)", "区域", "日期(yyyy-MM-dd)", "班次", "可用人数"])
    wb.save(path)


def main() -> None:
    data = json.loads(EXTRACTED.read_text(encoding="utf-8"))
    proc_by_mat, equip_by_gy, resources_meta = build_routing_index(data)
    cpe_routes = build_explicit_cpe_routes(data)

    report: dict[str, Any] = {
        "source": str(Path(r"d:\OneDrive\桌面\数据收集模板-盾安.xlsx")),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "orders": {"total": 0, "imported": 0, "issues": defaultdict(list)},
        "productResources": {"from_工艺_设备": 0, "from_CPE_关系": len(cpe_routes)},
        "unmatched": {},
        "warnings": [],
    }

    # ---------- product resources from 工艺 × 生产设备 ----------
    pr_seen: set[tuple[str, str, int | None]] = set()
    product_resources_out: list[dict] = []
    process_times: dict[str, list[float]] = defaultdict(list)

    for mat, keys in proc_by_mat.items():
        for key in keys:
            for eq in equip_by_gy.get(key, []):
                rid = eq["resourceId"]
                seq = safe_int(key[1], 1) if key[1] is not None else 1
                tup = (mat, rid, seq)
                if tup in pr_seen:
                    continue
                pr_seen.add(tup)
                product_resources_out.append({
                    "productCode": mat,
                    "resourceId": rid,
                    "setupTimeMinutes": 30,
                    "sequenceNo": seq,
                    "operationName": eq.get("operationName"),
                    "processTimeSeconds": 60.0,
                })
                process_times[rid].append(60.0)

    for route in cpe_routes:
        seq = route["sequenceNo"]
        tup = (route["productCode"], route["resourceId"], seq)
        if tup in pr_seen:
            continue
        pr_seen.add(tup)
        pt = route.get("processTimeSeconds") or 60.0
        product_resources_out.append({
            "productCode": route["productCode"],
            "resourceId": route["resourceId"],
            "setupTimeMinutes": 30,
            "sequenceNo": seq,
            "operationName": route.get("operationName"),
            "processTimeSeconds": pt,
        })
        process_times[route["resourceId"]].append(pt)

    report["productResources"]["from_工艺_设备"] = len(product_resources_out) - len(cpe_routes)
    report["productResources"]["explicit_cpe_routes"] = len(cpe_routes)

    # ---------- resources ----------
    resources_out: list[dict] = []
    for rid, meta in sorted(resources_meta.items()):
        times = process_times.get(rid, [])
        rate = 60.0
        if times:
            median_pt = statistics.median(times)
            if median_pt > 0:
                rate = max(1.0, round(3600 / median_pt, 2))
        resources_out.append({
            "resourceId": rid,
            "areaId": meta.get("areaId") or AREA_DEFAULT,
            "bottleneck": meta.get("bottleneck", False),
            "runRatePerHour": rate,
        })

    if resources_out:
        top = sorted(resources_out, key=lambda r: r["runRatePerHour"])[:3]
        for r in top:
            r["bottleneck"] = True

    resource_ids = {r["resourceId"] for r in resources_out}

    # ---------- sales orders ----------
    raw_orders: list[dict] = []
    line_no_by_order: dict[str, int] = defaultdict(int)
    for row in data["订单"]:
        mat = row.get("*产品代码")
        cpe_hint = row.get("*产品名称")
        qty = safe_float(row.get("*计划数量"), 0)
        due = parse_date(row.get("*交期"))
        if qty <= 0 or due is None:
            continue
        product_code, hits, issue = resolve_order_product(mat, cpe_hint, proc_by_mat, equip_by_gy)
        report["orders"]["total"] += 1
        if issue:
            report["orders"]["issues"][issue].append({
                "订单代码": row.get("订单代码"),
                "产品代码": mat,
                "产品名称列": cpe_hint,
                "计划号": row.get("*生产计划号"),
            })
        if product_code and hits:
            has_resource = any(h["resourceId"] in resource_ids for h in hits)
            if not has_resource:
                report["orders"]["issues"]["equipment_not_in_resources"].append({
                    "产品代码": product_code,
                    "订单代码": row.get("订单代码"),
                    "resources": [h["resourceId"] for h in hits[:3]],
                })
        if product_code is None:
            continue
        if product_code not in {p["productCode"] for p in product_resources_out}:
            report["orders"]["issues"]["no_product_resource"].append({
                "产品代码": product_code,
                "订单代码": row.get("订单代码"),
            })
        area = str(row.get("*车间代码") or AREA_DEFAULT)
        so_no = str(row.get("订单代码") or row.get("*生产计划号") or f"DO-{report['orders']['total']:04d}")
        line_no_by_order[so_no] += 10
        raw_orders.append({
            "salesOrderNo": so_no,
            "salesOrderLineNo": line_no_by_order[so_no],
            "customerCode": CUSTOMER,
            "productCode": product_code,
            "orderQty": qty,
            "dueDate": due,
            "areaId": area,
            "planWeek": row.get("*生产计划号"),
            "materialCode": mat,
            "cpeHint": cpe_hint,
        })
        report["orders"]["imported"] += 1

    if raw_orders:
        min_due = min(o["dueDate"] for o in raw_orders)
        target_min = date.today() + timedelta(days=1)
        offset = (target_min - min_due).days
    else:
        offset = 0

    sales_orders_out: list[dict] = []
    for o in raw_orders:
        shifted = o["dueDate"] + timedelta(days=offset)
        sales_orders_out.append({
            "salesOrderNo": str(o["salesOrderNo"]),
            "salesOrderLineNo": o["salesOrderLineNo"],
            "customerCode": o["customerCode"],
            "productCode": o["productCode"],
            "orderQty": o["orderQty"],
            "promiseDate": shifted.isoformat(),
            "dueDate": shifted.isoformat(),
            "priority": 5,
            "expediteLevel": 0,
            "status": "OPEN",
        })

    order_products = {o["productCode"] for o in sales_orders_out}
    pr_products = {p["productCode"] for p in product_resources_out}
    bom_scope = order_products | pr_products

    # ---------- BOM (only components for products in demand / routing) ----------
    bom_out: list[dict] = []
    bom_seen: set[tuple[str, str]] = set()
    for row in data["BOM（可选，用于物料齐套）"]:
        parent = row.get("产品代码")
        child = row.get("输出物品代码\n（子件物品代码）")
        if not parent or not child or parent == child:
            continue
        if str(parent).startswith("产品") or str(child).startswith("产品"):
            continue
        if parent not in bom_scope and child not in bom_scope:
            continue
        qty = safe_float(row.get("单位输出量\n（子件物品的用量）"), 1.0)
        base = safe_float(row.get("底数"), 1.0)
        if base > 0:
            qty = qty / base
        key = (parent, child)
        if key in bom_seen:
            continue
        bom_seen.add(key)
        bom_out.append({
            "parentProductCode": parent,
            "componentProductCode": child,
            "componentQty": round(qty, 6),
            "isCriticalComponent": True,
        })

    # ---------- lines (sample) ----------
    lines_out: list[dict] = []
    for res in resources_out[:12]:
        lines_out.append({
            "lineId": f"LINE-{res['resourceId'][:20]}",
            "areaId": res["areaId"],
            "resourceId": res["resourceId"],
            "lineMinHeadcount": 2,
            "lineCapacityPerShift": 480,
        })

    report["unmatched"] = {
        "summary": {
            k: len(v) for k, v in report["orders"]["issues"].items()
        },
        "notes": [
            "订单.*产品名称 列实际存的是 CPE 计划码（仅 3 行有值），多数订单仅提供物料号 *产品代码。",
            "产品工艺主键采用 工艺.产品代码（31 开头物料号），与 CPE 码通过 工艺路线+工序 间接关联。",
            "设备-产品生产关系 仅 8 行含 CPE 主数据，其余 8 万+ 行为物料展开，不单独导入。",
            "BOM 表为宽表+示例行混合，已过滤为与导入订单相关的父子件。",
            "模板无库存、换型矩阵；日历由系统启动时自动生成。",
        ],
    }
    report["orders"]["issues"] = {k: v[:50] if isinstance(v, list) else v for k, v in report["orders"]["issues"].items()}
    for k, v in report["unmatched"]["summary"].items():
        if v > 0:
            report["warnings"].append(f"订单 {k}: {v} 条")

    out_data = OrderedDict({
        "_meta": {
            "source": "盾安 数据收集模板-盾安.xlsx",
            "generated_at": report["generated_at"],
            "demand_date_offset_days": offset,
            "sales_order_count": len(sales_orders_out),
            "resource_count": len(resources_out),
            "product_resource_count": len(product_resources_out),
            "bom_count": len(bom_out),
            "workshop_code": AREA_DEFAULT,
        },
        "salesOrderLines": sales_orders_out,
        "bomComponents": bom_out,
        "inventory": [],
        "resources": resources_out,
        "productResources": product_resources_out,
        "lines": lines_out,
        "changeoverMatrix": [],
        "workOrders": [],
    })

    def build_lite(full: OrderedDict) -> OrderedDict:
        pr_keys = {(p["productCode"], p["resourceId"]) for p in full["productResources"]}
        ok_products = {p[0] for p in pr_keys}
        lite_orders = [o for o in full["salesOrderLines"] if o["productCode"] in ok_products]
        lite_products = {o["productCode"] for o in lite_orders}
        lite_pr = [p for p in full["productResources"] if p["productCode"] in lite_products]
        lite_res_ids = {p["resourceId"] for p in lite_pr}
        lite_resources = [r for r in full["resources"] if r["resourceId"] in lite_res_ids]
        lite_bom = [
            b for b in full["bomComponents"]
            if b["parentProductCode"] in lite_products or b["componentProductCode"] in lite_products
        ]
        meta = dict(full["_meta"])
        meta["variant"] = "lite"
        meta["sales_order_count"] = len(lite_orders)
        meta["product_resource_count"] = len(lite_pr)
        meta["bom_count"] = len(lite_bom)
        return OrderedDict({
            "_meta": meta,
            "salesOrderLines": lite_orders,
            "bomComponents": lite_bom,
            "inventory": [],
            "resources": lite_resources,
            "productResources": lite_pr,
            "lines": [ln for ln in full["lines"] if ln["resourceId"] in lite_res_ids][:12],
            "changeoverMatrix": [],
            "workOrders": [],
        })

    lite_data = build_lite(out_data)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_LITE.write_text(json.dumps(lite_data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_master_data_excel(lite_data, OUTPUT_EXCEL)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {OUTPUT}")
    print(f"Wrote {OUTPUT_LITE}")
    print(f"Wrote {OUTPUT_EXCEL}")
    print(f"Wrote {REPORT}")
    for k in ("salesOrderLines", "bomComponents", "resources", "productResources", "lines"):
        print(f"  {k}: {len(out_data[k])}")
    print("  order issues:", report["unmatched"]["summary"])


if __name__ == "__main__":
    main()
