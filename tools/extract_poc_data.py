"""Extract POC sheets into a single intermediate JSON for inspection."""
import json
from pathlib import Path

import openpyxl

PATH = Path(r"d:\OneDrive\工作\01 售前\03 已失败\202501_02_马勒#汽车零部件\POC需求资料.xlsx")
OUTPUT = Path(__file__).parent / "poc_extracted.json"

# Sheet column mapping: each entry lists Chinese-friendly key + English id + skip_rows from start.
SHEET_SPECS = {
    "物料基本资料": {
        "skip": 4,
        "fields": [
            (1, "part_id"),
            (2, "part_name"),
            (3, "safety_stock_qty"),
            (4, "unit"),
            (5, "part_type"),
            (6, "max_lot_size"),
            (7, "min_lot_size"),
            (8, "inc_lot_size"),
            (9, "is_phantom"),
            (10, "fixed_lead_time"),
            (11, "consolidate_day"),
            (13, "parent_part_id"),
            (14, "parent_qty"),
        ],
    },
    "物料清单": {
        "skip": 4,
        "fields": [
            (1, "output_part_id"),
            (2, "input_part_id"),
            (3, "sequence_txt"),
            (4, "output_part_qty"),
            (5, "input_part_qty"),
            (6, "is_phantom"),
        ],
    },
    "资源": {
        "skip": 5,
        "fields": [
            (1, "equip_id"),
            (2, "equip_name"),
            (3, "ws_id"),
            (4, "week_calendar_id"),
        ],
    },
    "资源群组": {
        "skip": 4,
        "fields": [
            (1, "equip_group_id"),
            (2, "equip_id"),
        ],
    },
    "工作中心": {
        "skip": 4,
        "fields": [
            (1, "ws_id"),
            (2, "ws_name"),
        ],
    },
    "工艺资料": {
        "skip": 5,
        "fields": [
            (1, "part_id"),
            (2, "route_id"),
            (3, "sequ_num"),
            (4, "operation_id"),
            (5, "operation_name"),
            (6, "equip_id"),
            (7, "equip_group_id"),
            (8, "setup_time"),
            (9, "process_time"),
            (10, "transfer_batch_size"),
        ],
    },
    "库存": {
        "skip": 4,
        "fields": [
            (0, "section"),
            (1, "part_id"),
            (2, "warehouse_id"),
            (3, "warehouse_name"),
            (4, "unallocate_qty"),
        ],
    },
    "需求来源": {
        "skip": 4,
        "fields": [
            (1, "demand_order_id"),
            (2, "customer_id"),
            (3, "due_date"),
            (4, "part_id"),
            (5, "part_name"),
            (6, "order_qty"),
            (7, "shipped_qty"),
            (8, "order_type"),
        ],
    },
}


def value(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True, read_only=True)
    output = {}
    for sheet_name, spec in SHEET_SPECS.items():
        if sheet_name not in wb.sheetnames:
            output[sheet_name] = []
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        non_empty = [r for r in all_rows if any(c is not None and str(c).strip() != "" for c in r)]
        body = non_empty[spec["skip"]:]
        records = []
        for row in body:
            rec = {}
            for col_idx, key in spec["fields"]:
                if col_idx < len(row):
                    rec[key] = value(row[col_idx])
                else:
                    rec[key] = None
            # Skip empty records
            if any(v not in (None, "") for v in rec.values()):
                records.append(rec)
        output[sheet_name] = records
    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    for name, recs in output.items():
        print(f"  {name}: {len(recs)} records")


if __name__ == "__main__":
    main()
