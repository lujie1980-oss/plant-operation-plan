#!/usr/bin/env python3
"""
从「MRP测试用例.xlsx」（晶华新材）生成 factory-jinghua-demo.json 与分切 Flyway 片段。

用法:
  python tools/parse_jinghua_mrp_excel.py
  python tools/parse_jinghua_mrp_excel.py --xlsx "d:/OneDrive/桌面/MRP测试用例.xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "tools" / "jinghua-mrp-source.xlsx"
OUT_JSON = ROOT / "src/main/resources/sample-data/factory-jinghua-demo.json"
OUT_SQL = ROOT / "src/main/resources/db/migration/V55__jinghua_mrp_slitting_refresh.sql"
ARCHIVE_XLSX = ROOT / "src/main/resources/sample-data/jinghua-mrp-test.xlsx"

HEADER_MARK = "母件料品_品名"
COL_PARENT_NAME, COL_PARENT_SPEC = 0, 1
COL_CHILD_NAME, COL_CHILD_SPEC = 2, 3
COL_OP = 4
COL_SO, COL_QTY, COL_UNIT, COL_DUE = 5, 6, 7, 8
COL_MAT_NAME, COL_MAT_SPEC, COL_INV_QTY, COL_INV_UNIT = 10, 11, 12, 13


def cell(row: list, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return str(v).strip()


def sql_str(s: str) -> str:
    return s.replace("'", "''")


def parse_due(text: str, base: date | None = None) -> str:
    base = base or date.today()
    if not text:
        return (base + timedelta(days=14)).isoformat()
    m = re.search(r"(\d+)", text)
    days = int(m.group(1)) if m else 14
    return (base + timedelta(days=days)).isoformat()


def parse_width_length_mm(spec: str) -> tuple[float | None, float | None]:
    if not spec:
        return None, None
    width = None
    length = None
    wm = re.search(r"(\d+)\s*mm", spec, re.I)
    if wm:
        width = float(wm.group(1))
    star = re.search(r"\*(\d+)\s*M", spec, re.I)
    if star:
        length = float(star.group(1)) * 1000.0
    else:
        tail = re.search(r"/(\d+)\s*M(?:/|$)", spec, re.I)
        if tail:
            length = float(tail.group(1)) * 1000.0
    return width, length


def roll_code_from_spec(spec: str) -> str:
    code = re.sub(r"[^\w\-/.]+", "-", spec)
    code = re.sub(r"-+", "-", code).strip("-")
    return f"MR-{code[:100]}" if code else "MR-UNKNOWN"


def order_code_from_spec(spec: str, idx: int) -> str:
    base = re.sub(r"[^\w\-/.]+", "-", spec)[:90]
    return f"CO-{base}-{idx}"


def resolve_child_code(child_name: str, child_spec: str) -> str:
    if child_spec:
        return child_spec
    if child_name:
        return child_name.strip()
    return ""


def is_pe_master(parent_name: str, parent_spec: str) -> bool:
    if not parent_spec:
        return False
    if "PE" in parent_name or "600M" in parent_spec:
        return True
    return "305" in parent_spec and "*" in parent_spec


def is_slitting_semi(spec: str) -> bool:
    return bool(spec and re.search(r"\d+\s*mm", spec, re.I))


def load_rows(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[list] = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row):
            rows.append(row)
    return rows


def add_bom_line(
    bom: list[dict],
    bom_seen: set[tuple[str, str, str]],
    finished: str,
    parent: str,
    child: str,
    *,
    critical: bool = True,
) -> None:
    if not finished or not parent or not child:
        return
    key = (finished, parent, child)
    if key in bom_seen:
        return
    bom_seen.add(key)
    bom.append(
        {
            "finishedProductCode": finished,
            "parentProductCode": parent,
            "componentProductCode": child,
            "componentQty": 1.0,
            "isCriticalComponent": critical,
        }
    )


def remap_bom_parent(code: str, from_fg: str, to_fg: str) -> str:
    return to_fg if code == from_fg else code


def consolidate_slitting_bom(bom: list[dict], root_fg: str) -> None:
    """分切演示只保留一条 BOM 树：去掉其它半成品根，子树归到成品根下。"""
    if not root_fg:
        return
    deduped: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for row in bom:
        finished = row["finishedProductCode"]
        parent = row["parentProductCode"]
        child = row["componentProductCode"]
        if finished != root_fg and parent == finished:
            continue
        normalized = {
            **row,
            "finishedProductCode": root_fg,
        }
        key = (parent, child)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(normalized)
    bom.clear()
    bom.extend(deduped)


def replicate_semi_bom(bom: list[dict], sales_products: list[str]) -> None:
    """将已解析的半成品 BOM 树复制到其它销售料号（Excel 通常只展开一条示例链）。"""
    by_finished: dict[str, list[dict]] = {}
    for row in bom:
        fg = row["finishedProductCode"]
        if is_slitting_semi(fg):
            by_finished.setdefault(fg, []).append(row)
    if not by_finished:
        return
    template_fg = max(by_finished.keys(), key=lambda k: len(by_finished[k]))
    template_rows = by_finished[template_fg]
    seen = {(r["finishedProductCode"], r["parentProductCode"], r["componentProductCode"]) for r in bom}
    for product in sales_products:
        if not is_slitting_semi(product) or product in by_finished:
            continue
        for row in template_rows:
            parent = remap_bom_parent(row["parentProductCode"], template_fg, product)
            key = (product, parent, row["componentProductCode"])
            if key in seen:
                continue
            seen.add(key)
            bom.append(
                {
                    "finishedProductCode": product,
                    "parentProductCode": parent,
                    "componentProductCode": row["componentProductCode"],
                    "componentQty": row["componentQty"],
                    "isCriticalComponent": row["isCriticalComponent"],
                }
            )


def collect_sales_products(rows: list[list]) -> list[str]:
    products: list[str] = []
    parent_spec = ""
    for row in rows:
        c = [cell(row, i) for i in range(16)]
        if c[COL_PARENT_SPEC]:
            parent_spec = c[COL_PARENT_SPEC]
        so = c[COL_SO]
        if so and re.match(r"^SO\d+$", so, re.I):
            product = c[COL_MAT_SPEC] or parent_spec or c[COL_CHILD_SPEC]
            if product and product not in products:
                products.append(product)
    return products


def parse_workbook(rows: list[list]) -> dict:
    sales_products = collect_sales_products(rows)
    parent_name = ""
    carry_parent = ""
    fg_pe = ""
    scope_finished = ""
    bom: list[dict] = []
    bom_seen: set[tuple[str, str, str]] = set()
    sales: list[dict] = []
    sales_seen: set[str] = set()
    inventory: dict[str, dict] = {}
    resources: dict[str, dict] = {}
    product_ops: dict[str, str] = {}

    for row in rows:
        c = [cell(row, i) for i in range(16)]
        if c[COL_PARENT_NAME] == HEADER_MARK:
            carry_parent = ""
            continue

        if c[COL_PARENT_NAME]:
            parent_name = c[COL_PARENT_NAME]
        if c[COL_PARENT_SPEC]:
            carry_parent = c[COL_PARENT_SPEC]
            if is_pe_master(parent_name, carry_parent):
                fg_pe = carry_parent
                scope_finished = carry_parent
            elif carry_parent in sales_products or (
                is_slitting_semi(carry_parent) and scope_finished in ("", fg_pe)
            ):
                scope_finished = carry_parent
            elif not scope_finished:
                scope_finished = carry_parent
            if c[COL_OP]:
                product_ops[carry_parent] = c[COL_OP]

        active_parent = carry_parent
        child_code = resolve_child_code(c[COL_CHILD_NAME], c[COL_CHILD_SPEC])
        if active_parent and child_code:
            finished = scope_finished or fg_pe or active_parent
            critical = not (
                child_code == c[COL_CHILD_NAME]
                and not c[COL_CHILD_SPEC]
                and c[COL_CHILD_NAME]
            ) or bool(c[COL_CHILD_SPEC])
            add_bom_line(bom, bom_seen, finished, active_parent, child_code, critical=critical)

        # 销售订单
        so = c[COL_SO]
        if so and re.match(r"^SO\d+$", so, re.I):
            so_key = so.upper()
            if so_key not in sales_seen:
                sales_seen.add(so_key)
                product = c[COL_MAT_SPEC] or parent_spec or child_spec
                qty_raw = c[COL_QTY]
                try:
                    qty = float(qty_raw) if qty_raw not in (None, "") else 1.0
                except (TypeError, ValueError):
                    qty = 1.0
                due = parse_due(c[COL_DUE])
                sales.append(
                    {
                        "salesOrderNo": so_key,
                        "salesOrderLineNo": 10,
                        "customerCode": "JINGHUA",
                        "productCode": product,
                        "orderQty": qty,
                        "promiseDate": due,
                        "dueDate": due,
                        "priority": 10 - len(sales),
                        "expediteLevel": 0,
                        "status": "OPEN",
                    }
                )
                if product and product not in sales_products:
                    sales_products.append(product)

        # 库存
        mat_spec = c[COL_MAT_SPEC]
        inv_qty = c[COL_INV_QTY]
        if mat_spec and inv_qty not in (None, ""):
            try:
                qty = float(inv_qty)
            except (TypeError, ValueError):
                continue
            prev = inventory.get(mat_spec)
            if prev is None or qty > prev["onhandQty"]:
                inventory[mat_spec] = {
                    "stockingPointCode": "WH-JINGHUA",
                    "productCode": mat_spec,
                    "onhandQty": qty,
                    "reservedQty": 0,
                }

        # 产线资源（案例资源区块）
        res_name = c[COL_MAT_NAME]
        if res_name in ("资源", "案例资源"):
            continue
        time_label = c[COL_MAT_SPEC]
        cap = c[COL_INV_QTY]
        if res_name and time_label and "H" in str(time_label):
            rid = res_name.replace("+", "-")
            if rid not in resources:
                try:
                    cap_n = int(float(cap)) if cap not in (None, "") else 2
                except (TypeError, ValueError):
                    cap_n = 2
                resources[rid] = {
                    "resourceId": rid,
                    "areaId": "美纹车间",
                    "bottleneck": rid in ("分条", "涂布", "含浸"),
                    "runRatePerHour": max(40, cap_n * 30),
                }

    consolidate_slitting_bom(bom, fg_pe)

    fg_parent_spec = fg_pe

    # 默认资源
    if not resources:
        for rid in ("分条", "开平", "涂布", "含浸", "离型"):
            resources[rid] = {
                "resourceId": rid,
                "areaId": "美纹车间",
                "bottleneck": rid == "分条",
                "runRatePerHour": 100,
            }

    resource_ids = list(resources.keys())
    default_res = resource_ids[0]

    product_resources: list[dict] = []
    pr_seen: set[tuple[str, str]] = set()
    all_products = set()
    for b in bom:
        all_products.add(b["parentProductCode"])
        all_products.add(b["componentProductCode"])
    for s in sales:
        all_products.add(s["productCode"])
    for inv in inventory:
        all_products.add(inv)

    for pcode in sorted(all_products):
        op = product_ops.get(pcode, "")
        rid = default_res
        for key, r in resources.items():
            if op and key in op:
                rid = r["resourceId"]
                break
        if "mm" in pcode and "分条" in resources:
            rid = resources["分条"]["resourceId"]
        key = (pcode, rid)
        if key not in pr_seen:
            pr_seen.add(key)
            product_resources.append(
                {
                    "productCode": pcode,
                    "resourceId": rid,
                    "setupTimeMinutes": 15,
                    "sequenceNo": 1,
                    "operationName": op or "生产",
                    "processTimeSeconds": 60,
                }
            )

    lines = [
        {
            "lineId": f"LINE-{r['resourceId']}",
            "areaId": r["areaId"],
            "resourceId": r["resourceId"],
            "lineMinHeadcount": 2,
            "lineCapacityPerShift": 480,
        }
        for r in resources.values()
    ]

    return {
        "bom": bom,
        "sales": sales,
        "inventory": list(inventory.values()),
        "resources": list(resources.values()),
        "product_resources": product_resources,
        "lines": lines,
        "fg_parent_spec": fg_parent_spec or parent_spec,
    }


def build_slitting_sql(parsed: dict, sales: list[dict]) -> str:
    """母卷 + 分切子订单（由半成品规格解析宽/长）。"""
    lines = [
        "-- 晶华 MRP 测试用例.xlsx → 分切主数据刷新（workspace: jinghua）",
        "-- 由 parse_jinghua_mrp_excel.py 生成；勿修改已应用的 V53",
        "",
    ]
    parent_spec = parsed.get("fg_parent_spec") or "M69/305*600M/1R/深黄"
    w, ln = parse_width_length_mm(parent_spec)
    w = w or 305.0
    ln = ln or 600000.0
    rc = roll_code_from_spec(parent_spec)
    fg = parsed.get("fg_parent_spec") or parent_spec
    lines.append(
        f"INSERT INTO master_roll (workspace_id, roll_code, width_mm, length_mm, thickness_mm, material_code, "
        f"product_code, finished_product_code, kerf_longitudinal_mm, kerf_transverse_mm, status)\n"
        f"SELECT 'jinghua', '{sql_str(rc)}', {w}, {ln}, NULL, 'PE', '{sql_str(fg)}', '{sql_str(fg)}', 2, 2, 'AVAILABLE'\n"
        f"WHERE NOT EXISTS (SELECT 1 FROM master_roll WHERE workspace_id = 'jinghua' AND roll_code = '{sql_str(rc)}');\n"
    )

    semi_specs: list[str] = []
    for inv in parsed["inventory"]:
        spec = inv["productCode"]
        if re.search(r"\d+\s*mm", spec, re.I):
            semi_specs.append(spec)
    for b in parsed["bom"]:
        child = b["componentProductCode"]
        if re.search(r"\d+\s*mm", child, re.I):
            semi_specs.append(child)

    seen_spec: set[str] = set()
    idx = 0
    so_by_product: dict[str, str] = {}
    for s in sales:
        so_by_product.setdefault(s["productCode"], s["salesOrderNo"])
    for spec in semi_specs:
        if spec in seen_spec:
            continue
        seen_spec.add(spec)
        cw, cl = parse_width_length_mm(spec)
        if not cw or not cl:
            continue
        idx += 1
        oc = order_code_from_spec(spec, idx)
        so = so_by_product.get(spec, "")
        so_clause = (
            f", '{sql_str(so)}', 10" if so else ", NULL, NULL"
        )
        fin = spec
        lines.append(
            f"INSERT INTO child_slitting_order (workspace_id, order_code, width_mm, length_mm, quantity, priority, "
            f"product_code, finished_product_code, sales_order_no, sales_order_line_no, status)\n"
            f"SELECT 'jinghua', '{sql_str(oc)}', {cw}, {cl}, 1, {20 - idx}, '{sql_str(spec)}', '{sql_str(fin)}'"
            f"{so_clause}, 'OPEN'\n"
            f"WHERE NOT EXISTS (SELECT 1 FROM child_slitting_order WHERE workspace_id = 'jinghua' AND order_code = '{sql_str(oc)}');\n"
        )

    lines.extend(
        [
            "",
            "UPDATE master_roll",
            "SET product_code = 'M69/305*600M/1R/深黄',",
            "    finished_product_code = 'M69/305*600M/1R/深黄'",
            "WHERE workspace_id = 'jinghua' AND roll_code = 'MR-M69/305-600M/1R/深黄'",
            "  AND (product_code IS NULL OR finished_product_code IS NULL);",
            "",
        ]
    )
    updates = [
        ("CO-M69/730mm/深黄/3M-1", "M69/730mm/深黄/3M"),
        ("CO-M69/730mm/浅黄/3M-2", "M69/730mm/浅黄/3M"),
        ("CO-M69/800mm/深黄/3M-3", "M69/800mm/深黄/3M"),
        ("CO-L80H/1515mm/深黄/3M-4", "L80H/1515mm/深黄/3M"),
        ("CO-GL60D/1520mm/深黄/3M-5", "GL60D/1520mm/深黄/3M"),
        ("CO-E48/1555mm/深黄/3M-6", "E48/1555mm/深黄/3M"),
    ]
    for oc, spec in updates:
        lines.append(
            f"UPDATE child_slitting_order SET product_code = '{sql_str(spec)}', "
            f"finished_product_code = '{sql_str(spec)}'\n"
            f"WHERE workspace_id = 'jinghua' AND order_code = '{sql_str(oc)}';\n"
        )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUT_JSON)
    parser.add_argument("--sql", type=Path, default=OUT_SQL)
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"Excel 不存在: {args.xlsx}")

    rows = load_rows(args.xlsx)
    parsed = parse_workbook(rows)

    demo = {
        "_meta": {
            "source": "MRP测试用例.xlsx（晶华新材）",
            "source_path": str(args.xlsx.name),
            "generated_at": date.today().isoformat(),
            "sales_order_count": len(parsed["sales"]),
            "bom_count": len(parsed["bom"]),
            "resource_count": len(parsed["resources"]),
            "product_resource_count": len(parsed["product_resources"]),
            "work_centers": [{"ws_id": "美纹车间", "ws_name": "JH"}],
        },
        "salesOrderLines": parsed["sales"],
        "bomComponents": parsed["bom"],
        "inventory": parsed["inventory"],
        "resources": parsed["resources"],
        "productResources": parsed["product_resources"],
        "lines": parsed["lines"],
        "changeoverMatrix": [],
        "workOrders": [],
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(demo, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.out}")
    fg_counts: dict[str, int] = {}
    for b in parsed["bom"]:
        fg = b.get("finishedProductCode", "?")
        fg_counts[fg] = fg_counts.get(fg, 0) + 1
    print(f"  sales={len(parsed['sales'])} bom={len(parsed['bom'])} inv={len(parsed['inventory'])} res={len(parsed['resources'])}")
    for fg, n in sorted(fg_counts.items(), key=lambda x: (-x[1], x[0])):
        print(f"    bom[{fg}]={n}")

    sql = build_slitting_sql(parsed, parsed["sales"])
    args.sql.write_text(sql, encoding="utf-8")
    print(f"Wrote {args.sql}")

    if args.xlsx != ARCHIVE_XLSX:
        ARCHIVE_XLSX.parent.mkdir(parents=True, exist_ok=True)
        import shutil

        shutil.copy2(args.xlsx, ARCHIVE_XLSX)
        print(f"Archived xlsx -> {ARCHIVE_XLSX}")


if __name__ == "__main__":
    main()
