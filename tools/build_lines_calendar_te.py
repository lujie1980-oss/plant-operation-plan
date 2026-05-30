"""Build 产线 + 资源日历 Excel from te routing subset and import via API."""
import json
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROUTING_SUBSET = Path(__file__).resolve().parent.parent / ".import-tmp" / "subset-routing-bom-100.xlsx"
OUT_XLSX = Path(__file__).resolve().parent.parent / ".import-tmp" / "te-lines-calendar.xlsx"
HORIZON_DAYS = 90
SHIFT_ID = "DAY"
AVAILABLE_MINUTES = 480
DEFAULT_AREA = "1294"
DEFAULT_MIN_HEADCOUNT = 2
DEFAULT_LINE_CAPACITY = 480


def read_header(ws):
    header = []
    for c in range(1, 40):
        v = ws.cell(1, c).value
        if v is None and c > 1 and not header:
            break
        if v is None and header:
            break
        header.append(str(v).strip() if v is not None else "")
    while header and header[-1] == "":
        header.pop()
    return header


def cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def collect_lines_from_routing(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = read_header(ws)
    idx = {h: i + 1 for i, h in enumerate(header)}

    lines: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        resource_id = cell_str(ws.cell(r, idx["设备组"]).value)
        if not resource_id:
            continue
        area_col = idx.get("厂区代码")
        area = cell_str(ws.cell(r, area_col).value) if area_col else DEFAULT_AREA
        if not area:
            area = DEFAULT_AREA
        line_col = idx.get("工序线体")
        line_id = cell_str(ws.cell(r, line_col).value) if line_col else resource_id
        if not line_id:
            line_id = resource_id
        if line_id not in lines:
            lines[line_id] = {
                "lineId": line_id,
                "areaId": area,
                "resourceId": resource_id,
                "lineMinHeadcount": DEFAULT_MIN_HEADCOUNT,
                "lineCapacityPerShift": DEFAULT_LINE_CAPACITY,
            }
    return list(lines.values())


def collect_resources_from_routing(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header = read_header(ws)
    idx = {h: i + 1 for i, h in enumerate(header)}
    resources: set[str] = set()
    for r in range(2, ws.max_row + 1):
        rid = cell_str(ws.cell(r, idx["设备组"]).value)
        if rid:
            resources.add(rid)
    return sorted(resources)


def build_calendar_rows(resource_ids: list[str], start: date, days: int) -> list[dict]:
    rows = []
    for rid in resource_ids:
        for d in range(days):
            cal_date = start + timedelta(days=d)
            rows.append(
                {
                    "resourceId": rid,
                    "shiftId": SHIFT_ID,
                    "calendarDate": cal_date.isoformat(),
                    "availableCapacityMinutes": AVAILABLE_MINUTES,
                    "unavailableCapacityMinutes": 0,
                }
            )
    return rows


def write_workbook(lines: list[dict], calendars: list[dict], out: Path) -> None:
    """Column order must match MasterDataExcelSheet (by index, not header text)."""
    wb = Workbook()
    ws_line = wb.active
    ws_line.title = "产线"
    line_cols = [
        ("id", None),
        ("lineId", "lineId"),
        ("areaId", "areaId"),
        ("resourceId", "resourceId"),
        ("lineMinHeadcount", "lineMinHeadcount"),
        ("lineCapacityPerShift", "lineCapacityPerShift"),
    ]
    for c, (hdr, _) in enumerate(line_cols, start=1):
        ws_line.cell(1, c).value = hdr
    for r, row in enumerate(lines, start=2):
        for c, (_, key) in enumerate(line_cols, start=1):
            ws_line.cell(r, c).value = "" if key is None else row[key]

    ws_cal = wb.create_sheet("资源日历")
    cal_cols = [
        ("id", None),
        ("resourceId", "resourceId"),
        ("calendarDate", "calendarDate"),
        ("shiftId", "shiftId"),
        ("availableCapacityMinutes", "availableCapacityMinutes"),
        ("unavailableCapacityMinutes", "unavailableCapacityMinutes"),
    ]
    for c, (hdr, _) in enumerate(cal_cols, start=1):
        ws_cal.cell(1, c).value = hdr
    for r, row in enumerate(calendars, start=2):
        for c, (_, key) in enumerate(cal_cols, start=1):
            ws_cal.cell(r, c).value = "" if key is None else row[key]

    wb.save(out)


def main():
    if not ROUTING_SUBSET.exists():
        raise FileNotFoundError(f"Missing routing subset: {ROUTING_SUBSET}")

    lines = collect_lines_from_routing(ROUTING_SUBSET)
    resources = collect_resources_from_routing(ROUTING_SUBSET)
    start = date.today()
    calendars = build_calendar_rows(resources, start, HORIZON_DAYS)
    write_workbook(lines, calendars, OUT_XLSX)

    print(
        json.dumps(
            {
                "lines": len(lines),
                "resources": len(resources),
                "calendar_rows": len(calendars),
                "horizon_days": HORIZON_DAYS,
                "start_date": start.isoformat(),
                "output": str(OUT_XLSX),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
