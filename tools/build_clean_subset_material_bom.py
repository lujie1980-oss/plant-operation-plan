from pathlib import Path
from openpyxl import load_workbook, Workbook

src = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\routing-bom.xlsx")
out = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\subset-material-bom-100-clean.xlsx")

wb = load_workbook(src, data_only=True)
ws = wb[wb.sheetnames[0]]

# header row detect
header_row = 1
header = [ws.cell(header_row, c).value for c in range(1, 80)]
header = [str(v).strip() if v is not None else "" for v in header]
while header and header[-1] == "":
    header.pop()
idx = {h:i for i,h in enumerate(header)}

finished_i = idx["成品料号"] + 1

# pick first 100 finished codes
top = []
seen = set()
for r in range(header_row+1, ws.max_row+1):
    v = ws.cell(r, finished_i).value
    s = "" if v is None else str(v).strip()
    if not s or s in seen:
        continue
    seen.add(s)
    top.append(s)
    if len(top) >= 100:
        break

top_set = set(top)

keep_cols = [
    "基地代码", "BOM版本", "BOM生效时间", "BOM失效时间", "成品料号", "产品代码",
    "组件代码", "不计算齐套率", "组件数量", "组件损耗率", "组件生效时间", "组件失效时间"
]
keep_idx = [idx[c] + 1 for c in keep_cols if c in idx]
keep_headers = [c for c in keep_cols if c in idx]

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Sheet1"

for c,h in enumerate(keep_headers, start=1):
    out_ws.cell(1,c).value = h

out_r = 2
for r in range(header_row+1, ws.max_row+1):
    finished = ws.cell(r, finished_i).value
    fs = "" if finished is None else str(finished).strip()
    if fs not in top_set:
        continue
    for c,src_c in enumerate(keep_idx, start=1):
        v = ws.cell(r, src_c).value
        out_ws.cell(out_r, c).value = v
    out_r += 1

out_wb.save(out)
print({"rows": out_r-2, "file": str(out)})
