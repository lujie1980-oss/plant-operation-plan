"""Recreate workspace te and import 100-finished subset master data + lines/calendar."""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path
from urllib import error, request

ROOT = Path(__file__).resolve().parent.parent
TMP = ROOT / ".import-tmp"
BASE = "http://localhost:8080/api/v1"
WORKSPACE = "te"


def find_template_dir() -> Path:
    for base in (Path(r"d:\OneDrive\桌面"), Path.home() / "OneDrive" / "桌面"):
        if not base.is_dir():
            continue
        for sub in base.iterdir():
            if sub.is_dir() and list(sub.glob("*BOM*.xlsx")):
                return sub
    raise FileNotFoundError("Cannot find 数据模板 folder with *BOM*.xlsx")


def _is_real_xlsx(path: Path) -> bool:
    if path.name.startswith("~$") or path.stat().st_size < 100_000:
        return False
    return path.read_bytes()[:2] == b"PK"


def _sheet_headers(path: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=40, values_only=True))
    wb.close()
    return [str(v).strip() if v is not None else "" for v in row if v not in (None, "")]


def copy_sources(tpl_dir: Path) -> None:
    TMP.mkdir(parents=True, exist_ok=True)
    xlsx = [p for p in tpl_dir.glob("*.xlsx") if _is_real_xlsx(p)]
    bom_candidates = [p for p in xlsx if p.stat().st_size > 1_000_000]
    material_src = routing_src = None
    for p in bom_candidates:
        headers = _sheet_headers(p)
        if "组件代码" in headers and "成品料号" in headers:
            material_src = p
        if "设备组" in headers and ("工序编号" in headers or "制造CT" in headers):
            routing_src = p
    if material_src is None or routing_src is None:
        raise FileNotFoundError(
            f"Could not identify material/routing BOM by header in {tpl_dir}"
        )
    # Names follow build_100_finished_subsets.py convention (routing-bom = 物料 BOM file)
    shutil.copy2(material_src, TMP / "routing-bom.xlsx")
    shutil.copy2(routing_src, TMP / "material-bom.xlsx")
    for f in xlsx:
        if 350_000 < f.stat().st_size < 450_000:
            shutil.copy2(f, TMP / "material-master-source.xlsx")
            break


def post_json(path: str, payload: dict, workspace: str | None = WORKSPACE) -> dict:
    headers = {"Content-Type": "application/json"}
    if workspace:
        headers["X-Workspace-Id"] = workspace
    req = request.Request(
        f"{BASE}{path}",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers=headers,
    )
    with request.urlopen(req, timeout=300) as resp:
        return json.loads(resp.read().decode("utf-8"))


def upload_excel(path: Path, replace: bool = True) -> dict:
    data = path.read_bytes()
    url = f"{BASE}/master-data/excel/import?replace={'true' if replace else 'false'}"
    req = request.Request(
        url,
        data=data,
        method="POST",
        headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "X-Workspace-Id": WORKSPACE,
        },
    )
    try:
        with request.urlopen(req, timeout=600) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as e:
        body = e.read().decode("utf-8", "ignore")
        raise RuntimeError(f"Import failed {path.name}: {e.code} {body[:500]}") from e


def run_py(script: str) -> None:
    subprocess.run([sys.executable, str(ROOT / "tools" / script)], check=True, cwd=ROOT)


def ensure_workspace() -> None:
    try:
        workspaces = json.loads(
            request.urlopen(f"{BASE}/workspaces", timeout=30).read().decode("utf-8")
        )
        if isinstance(workspaces, dict) and "value" in workspaces:
            workspaces = workspaces["value"]
        if any(w.get("workspaceId") == WORKSPACE for w in workspaces):
            return
    except Exception:
        pass
    post_json("/workspaces", {"id": WORKSPACE, "name": "TE"}, workspace=None)


def count(endpoint: str) -> int:
    req = request.Request(
        f"{BASE}/master-data/{endpoint}",
        headers={"X-Workspace-Id": WORKSPACE},
    )
    with request.urlopen(req, timeout=120) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return len(data) if isinstance(data, list) else 0


def main() -> None:
    tpl = find_template_dir()
    copy_sources(tpl)

    run_py("build_100_finished_subsets.py")
    run_py("build_clean_subset_material_bom.py")
    run_py("normalize_subset_bom_dates.py")
    run_py("finalize_subset_bom_dates.py")
    run_py("build_subset_material_master_100.py")
    run_py("build_lines_calendar_te.py")

    ensure_workspace()

    results = {}
    for key, fname in [
        ("materials", "subset-material-master-100.xlsx"),
        ("boms", "subset-material-bom-100-final.xlsx"),
        ("routing", "subset-routing-bom-100.xlsx"),
    ]:
        p = TMP / fname
        r = upload_excel(p, replace=True)
        results[key] = r

    run_py("import_lines_calendar_api_te.py")

    summary = {
        "workspace": WORKSPACE,
        "imports": results,
        "counts": {
            "materials": count("materials"),
            "boms": count("boms"),
            "product-resources": count("product-resources"),
            "resources": count("resources"),
            "lines": count("lines"),
            "calendar": count("calendar"),
        },
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
