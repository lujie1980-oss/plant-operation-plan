from io import BytesIO
from urllib import request, error
import json
from openpyxl import load_workbook, Workbook

SRC = r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\subset-material-bom-100-clean.xlsx"
ws = load_workbook(SRC, data_only=True).active
header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
while header and header[-1] in (None, ""):
    header.pop()
vals = [ws.cell(2, c).value for c in range(1, len(header) + 1)]

one = Workbook(); out = one.active; out.title='Sheet1'
for c,h in enumerate(header,1):
    out.cell(1,c).value=h
    out.cell(2,c).value=vals[c-1]
bio=BytesIO(); one.save(bio)
req=request.Request('http://localhost:8080/api/v1/master-data/excel/import',data=bio.getvalue(),method='POST')
req.add_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
req.add_header('X-Workspace-Id','te')
try:
    with request.urlopen(req, timeout=60) as resp:
        print(resp.status, resp.read().decode('utf-8','ignore'))
except error.HTTPError as e:
    print('status', e.code)
    print(e.read().decode('utf-8','ignore'))
