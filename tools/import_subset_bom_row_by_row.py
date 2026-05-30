from io import BytesIO
from urllib import request, error
import json
from openpyxl import load_workbook, Workbook

SRC = r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\subset-material-bom-100-clean.xlsx"
URL = "http://localhost:8080/api/v1/master-data/excel/import"
WORKSPACE = "te"

wb = load_workbook(SRC, data_only=True)
ws = wb[wb.sheetnames[0]]
header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
while header and header[-1] in (None, ""):
    header.pop()
cols = len(header)

success = 0
failed = 0
samples = []

for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, cols + 1)]
    if all(v in (None, "") for v in vals):
        continue

    one = Workbook()
    out = one.active
    out.title = "Sheet1"
    for c, h in enumerate(header, start=1):
        out.cell(1, c).value = h
        out.cell(2, c).value = vals[c - 1]

    bio = BytesIO()
    one.save(bio)
    data = bio.getvalue()

    req = request.Request(URL, data=data, method="POST")
    req.add_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    req.add_header("X-Workspace-Id", WORKSPACE)

    try:
        with request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8", errors="ignore")
        j = json.loads(body) if body else {}
        if j.get("errors"):
            failed += 1
            if len(samples) < 10:
                samples.append({"row": r, "errors": j.get("errors")[:2]})
        else:
            success += 1
    except Exception as ex:
        failed += 1
        if len(samples) < 10:
            samples.append({"row": r, "error": str(ex)})

print(json.dumps({"success_rows": success, "failed_rows": failed, "samples": samples}, ensure_ascii=False))
