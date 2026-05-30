"""Extract 盾安数据收集模板 into structured JSON for conversion."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

SOURCE = Path(r"d:\OneDrive\桌面\数据收集模板-盾安.xlsx")
OUTPUT = Path(__file__).parent / "dunan_extracted.json"


def norm(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip().replace("\t", "")
    return s or None


NO_FORWARD_FILL: dict[str, set[str]] = {
    "订单": {"*产品名称", "*生产计划号", "订单代码", "*产品代码", "*计划数量", "*交期"},
    "设备-产品生产关系": {"*产品代码", "*产品名称", "*设备代码", "*设备名称", "*工序类型", "步骤号", "*批量", "*单批次生产时长"},
}


def rows_with_forward_fill(ws, data_start: int = 3, sheet_name: str = "") -> list[dict[str, Any]]:
    """Parse sheet rows; duplicate Chinese headers get _2, _3 suffixes."""
    raw = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, r in enumerate(raw):
        if any(c and "*" in str(c) for c in r):
            header_row = i
            break
        if any(c and str(c).strip() in ("产品代码", "设备组") for c in r):
            header_row = i
            break
    if header_row is None:
        return []

    header_counts: dict[str, int] = {}
    fields: list[str] = []
    for c in raw[header_row]:
        base = norm(c) or "col"
        header_counts[base] = header_counts.get(base, 0) + 1
        suffix = f"_{header_counts[base]}" if header_counts[base] > 1 else ""
        fields.append(base + suffix)

    skip_ff = NO_FORWARD_FILL.get(sheet_name, set())
    out: list[dict[str, Any]] = []
    carry: dict[str, str | None] = {}
    for r in raw[header_row + 1 :]:
        if all(norm(c) is None for c in r):
            continue
        row: dict[str, Any] = {}
        for i, field in enumerate(fields):
            val = norm(r[i]) if i < len(r) else None
            if val is not None and field not in skip_ff:
                carry[field] = val
            row[field] = val if field in skip_ff else carry.get(field)
        if any(v is not None for v in row.values()):
            out.append(row)
    return out


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    data = {name: rows_with_forward_fill(wb[name], sheet_name=name) for name in wb.sheetnames}
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    for name, rows in data.items():
        print(f"  {name}: {len(rows)} rows")


if __name__ == "__main__":
    main()
