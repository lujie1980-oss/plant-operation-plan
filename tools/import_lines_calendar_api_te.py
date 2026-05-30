"""Import 产线 + 资源日历 into te via REST API (skips full-workbook validation)."""
import json
from pathlib import Path
from urllib import request, error

BASE = "http://localhost:8080/api/v1"
WORKSPACE = "te"
XLSX = Path(__file__).resolve().parent.parent / ".import-tmp" / "te-lines-calendar.xlsx"


def post_json(path: str, payload: dict) -> dict:
    body = json.dumps(payload).encode("utf-8")
    req = request.Request(
        f"{BASE}{path}",
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "X-Workspace-Id": WORKSPACE,
        },
    )
    with request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))


def main():
    from openpyxl import load_workbook

    wb = load_workbook(XLSX, data_only=True)
    lines_ws = wb["产线"]
    cal_ws = wb["资源日历"]

    line_ok = 0
    line_fail = 0
    for r in range(2, lines_ws.max_row + 1):
        line_id = lines_ws.cell(r, 2).value
        if line_id is None or str(line_id).strip() == "":
            continue
        payload = {
            "lineId": str(line_id).strip(),
            "areaId": str(lines_ws.cell(r, 3).value or "1294").strip(),
            "resourceId": str(lines_ws.cell(r, 4).value).strip(),
            "lineMinHeadcount": int(float(lines_ws.cell(r, 5).value or 2)),
            "lineCapacityPerShift": int(float(lines_ws.cell(r, 6).value or 480)),
        }
        try:
            post_json("/master-data/lines", payload)
            line_ok += 1
        except error.HTTPError as e:
            line_fail += 1
            if line_fail <= 3:
                print("line fail", payload["lineId"], e.read().decode("utf-8", "ignore")[:200])

    cal_ok = 0
    cal_fail = 0
    for r in range(2, cal_ws.max_row + 1):
        resource_id = cal_ws.cell(r, 2).value
        if resource_id is None or str(resource_id).strip() == "":
            continue
        cal_date = cal_ws.cell(r, 3).value
        if hasattr(cal_date, "date"):
            cal_date = cal_date.date().isoformat()
        else:
            cal_date = str(cal_date).strip()[:10]
        payload = {
            "resourceId": str(resource_id).strip(),
            "shiftId": str(cal_ws.cell(r, 4).value or "DAY").strip(),
            "calendarDate": cal_date,
            "availableCapacityMinutes": int(float(cal_ws.cell(r, 5).value or 480)),
            "unavailableCapacityMinutes": int(float(cal_ws.cell(r, 6).value or 0)),
        }
        try:
            post_json("/master-data/calendar", payload)
            cal_ok += 1
        except error.HTTPError as e:
            cal_fail += 1
            if cal_fail <= 3:
                print("cal fail", payload["resourceId"], cal_date, e.read().decode("utf-8", "ignore")[:200])

    print(
        json.dumps(
            {"lines_ok": line_ok, "lines_fail": line_fail, "calendar_ok": cal_ok, "calendar_fail": cal_fail},
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
