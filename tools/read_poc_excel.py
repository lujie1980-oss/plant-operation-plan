"""Inspect the Mahle POC Excel workbook structure: sheets, headers, row count, sample rows."""
import json
from pathlib import Path

import openpyxl

PATH = Path(r"d:\OneDrive\工作\01 售前\03 已失败\202501_02_马勒#汽车零部件\POC需求资料.xlsx")
OUTPUT = Path(__file__).parent / "poc_overview.json"


def main() -> None:
    wb = openpyxl.load_workbook(PATH, data_only=True, read_only=True)
    overview = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        header = non_empty[0] if non_empty else []
        sample = non_empty[1:8]
        overview[name] = {
            "row_count": len(non_empty),
            "header": list(header),
            "sample_rows": [list(r) for r in sample],
        }
    OUTPUT.write_text(json.dumps(overview, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
