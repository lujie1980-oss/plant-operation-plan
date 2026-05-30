"""Build material master subset for the same 100 finished products as BOM subset."""
import json
from pathlib import Path

from openpyxl import load_workbook, Workbook


def read_header(ws):
    for r in range(1, 6):
        vals = [ws.cell(r, c).value for c in range(1, 80)]
        if any(v not in (None, "") for v in vals):
            header = [str(v).strip() if v is not None else "" for v in vals]
            while header and header[-1] == "":
                header.pop()
            return r, header
    raise ValueError(f"No header found in sheet {ws.title}")


def cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def find_material_master_file(one_drive_tpl: Path) -> Path:
    for f in one_drive_tpl.glob("*.xlsx"):
        if 350_000 < f.stat().st_size < 450_000:
            return f
    raise FileNotFoundError("Material master template not found by size")


def collect_codes_from_bom(bom_path: Path) -> set[str]:
    wb = load_workbook(bom_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, header = read_header(ws)
    idx = {h: i for i, h in enumerate(header)}
    codes: set[str] = set()
    for key in ("成品料号", "产品代码", "组件代码"):
        if key not in idx:
            continue
        col = idx[key]
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c + 1).value for c in range(len(header))]
            if all(v in (None, "") for v in row):
                continue
            s = cell_str(row[col])
            if s:
                codes.add(s)
    return codes


def main():
    root = Path(__file__).resolve().parent.parent
    tmp = root / ".import-tmp"
    bom_subset = tmp / "subset-material-bom-100-final.xlsx"
    if not bom_subset.exists():
        bom_subset = tmp / "subset-material-bom-100-normalized.xlsx"
    if not bom_subset.exists():
        raise FileNotFoundError("BOM subset file missing; run build_100_finished_subsets first")

    # OneDrive template dir (same discovery as import scripts)
    one_drive = Path.home() / "OneDrive"
    desktop_dirs = list(one_drive.glob("*"))
    tpl_dir = None
    for d in desktop_dirs:
        if not d.is_dir():
            continue
        for sub in d.iterdir():
            if sub.is_dir() and any(sub.glob("*BOM*.xlsx")):
                tpl_dir = sub
                break
        if tpl_dir:
            break
    local_master = tmp / "material-master-source.xlsx"
    if local_master.exists():
        master_src = local_master
    elif tpl_dir is not None:
        master_src = find_material_master_file(tpl_dir)
    else:
        raise FileNotFoundError(
            "Cannot locate material master source xlsx; copy 物料主数据模板.xlsx to .import-tmp/material-master-source.xlsx"
        )

    needed = collect_codes_from_bom(bom_subset)
    out_path = tmp / "subset-material-master-100.xlsx"

    wb = load_workbook(master_src, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, header = read_header(ws)
    if "产品代码" not in header:
        raise ValueError("Material master missing 产品代码 column")
    code_col = header.index("产品代码")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    for c, h in enumerate(header, start=1):
        out_ws.cell(1, c).value = h

    out_r = 2
    seen = set()
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c + 1).value for c in range(len(header))]
        if all(v in (None, "") for v in row_vals):
            continue
        code = cell_str(row_vals[code_col])
        if not code or code not in needed:
            continue
        if code in seen:
            continue
        seen.add(code)
        for c, v in enumerate(row_vals, start=1):
            out_ws.cell(out_r, c).value = v
        out_r += 1

    out_wb.save(out_path)
    missing = sorted(needed - seen)
    print(
        json.dumps(
            {
                "material_master_source": str(master_src),
                "bom_subset": str(bom_subset),
                "needed_codes": len(needed),
                "imported_rows": out_r - 2,
                "missing_codes": len(missing),
                "missing_sample": missing[:20],
                "output": str(out_path),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
