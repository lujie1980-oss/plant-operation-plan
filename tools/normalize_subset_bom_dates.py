from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from pathlib import Path

src = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\subset-material-bom-100-clean.xlsx")
out = Path(r"d:\AILab\PlantOperationPlan\plant-operation-plan\.import-tmp\subset-material-bom-100-normalized.xlsx")

wb = load_workbook(src, data_only=True)
ws = wb.active
headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
while headers and headers[-1] in (None,""):
    headers.pop()
cols = len(headers)
idx = {str(h).strip(): i+1 for i,h in enumerate(headers)}
date_cols = [idx[k] for k in ["BOM生效时间","BOM失效时间","组件生效时间","组件失效时间"] if k in idx]

out_wb = Workbook(); out_ws = out_wb.active; out_ws.title='Sheet1'
for c,h in enumerate(headers,1):
    out_ws.cell(1,c).value = h

out_r = 2
for r in range(2, ws.max_row+1):
    vals = [ws.cell(r,c).value for c in range(1, cols+1)]
    if all(v in (None,"") for v in vals):
        continue
    for c,v in enumerate(vals,1):
        if c in date_cols and v not in (None,""):
            if isinstance(v, datetime):
                v = v.date().isoformat()
            elif isinstance(v, date):
                v = v.isoformat()
            else:
                s = str(v).strip()
                if " " in s:
                    s = s.split(" ",1)[0]
                v = s.replace("/", "-")
        out_ws.cell(out_r,c).value = v
    out_r += 1

out_wb.save(out)
print({"rows": out_r-2, "file": str(out)})
